import openpyxl
import random
import string

# Function to generate a unique passcode
def generate_passcode(length=12):
    characters = string.ascii_uppercase + string.ascii_lowercase + string.digits
    passcode = ''.join(random.choice(characters) for _ in range(length))
    return passcode

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Define the number of passcodes per section
num_passcodes_per_section = 100
column_width = 20  # Width of columns to fit the passcodes

# Loop through the alphabet to create 26 sheets
for letter in string.ascii_uppercase:
    sheet = workbook.create_sheet(title=letter)
    
    # Loop through the alphabet to create columns A to Z in each sheet
    for col_idx, col_letter in enumerate(string.ascii_uppercase, start=1):
        # Set the column width
        sheet.column_dimensions[col_letter].width = column_width

        # Generate unique passcodes for the column
        passcodes = set()
        for row in range(1, num_passcodes_per_section + 1):
            passcode = generate_passcode()
            # Ensure passcode uniqueness
            while passcode in passcodes:
                passcode = generate_passcode()
            passcodes.add(passcode)
            sheet.cell(row=row, column=col_idx, value=passcode)

# Remove the default sheet created by openpyxl
if 'Sheet' in workbook.sheetnames:
    del workbook['Sheet']

# Save the workbook
workbook.save('UniquePasscodes.xlsx')

print("Excel file with multiple sheets and unique passcodes created successfully.")
